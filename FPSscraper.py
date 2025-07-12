import requests
from bs4 import BeautifulSoup
from requests.exceptions import ConnectTimeout, ConnectionError,ReadTimeout
from urllib3.exceptions import ProtocolError
import time
import re
from openpyxl import Workbook, load_workbook
import os

def format_page(soup):
    text = "\n" + "------------------------------------------" + "\n"
    for person in soup:
        elements = person.find_all('a', href=True)
        for i in elements:
            text += i.text + "\n"

        text += i.text + "\n" + "------------------------------------------" + "\n"
    return text

def get_state_code(state_name):
    state_mapping = {
        "Alabama": "AL", "Alaska": "AK", "Arizona": "AZ", "Arkansas": "AR", "California": "CA",
        "Colorado": "CO", "Connecticut": "CT", "Delaware": "DE", "Florida": "FL", "Georgia": "GA",
        "Hawaii": "HI", "Idaho": "ID", "Illinois": "IL", "Indiana": "IN", "Iowa": "IA",
        "Kansas": "KS", "Kentucky": "KY", "Louisiana": "LA", "Maine": "ME", "Maryland": "MD",
        "Massachusetts": "MA", "Michigan": "MI", "Minnesota": "MN", "Mississippi": "MS", "Missouri": "MO",
        "Montana": "MT", "Nebraska": "NE", "Nevada": "NV", "New Hampshire": "NH", "New Jersey": "NJ",
        "New Mexico": "NM", "New York": "NY", "North Carolina": "NC", "North Dakota": "ND", "Ohio": "OH",
        "Oklahoma": "OK", "Oregon": "OR", "Pennsylvania": "PA", "Rhode Island": "RI", "South Carolina": "SC",
        "South Dakota": "SD", "Tennessee": "TN", "Texas": "TX", "Utah": "UT", "Vermont": "VT",
        "Virginia": "VA", "Washington": "WA", "West Virginia": "WV", "Wisconsin": "WI", "Wyoming": "WY"
    }

    return state_mapping.get(state_name.title(), None)
def split_address(address):
    if "-" in address:
        address=address.split("-")[1]
    street_pattern = re.compile(r'(\d+)\s+(.+)')
    temp=address.split(",")
    street_match = street_pattern.match(temp[0])
    if street_match:
        number, street = street_match.groups()
        city = temp[1].strip()  # Remove any extra spaces
        state=temp[2].strip()
        if len(state)!=2:
            state = get_state_code(state)  # Remove any extra spaces
        zip_code = temp[3].strip()  # Remove any extra spaces
        return number,street,city,state,zip_code
    return None
def get_address(text):
    # Split the text into lines and remove leading/trailing spaces
    lines = text.strip().splitlines()

    # Find the first non-empty line
    first_non_empty_index = None
    for i, line in enumerate(lines):
        if line.strip():  # Found the first non-empty line
            first_non_empty_index = i
            break

    # If there's no valid first non-empty line, return None
    if first_non_empty_index is None:
        return None

    # Collect the next two non-empty lines
    next_two_lines = []
    for line in lines[first_non_empty_index + 1:]:
        if line.strip():  # Skip empty lines
            next_two_lines.append(line.strip())
        if len(next_two_lines) == 2:  # Stop after collecting two lines
            break

    # Join the two lines with a comma, if available
    return ', '.join(next_two_lines) if len(next_two_lines) == 2 else None


def extract_phone_numbers(text):
    # Regular expression pattern to match US phone numbers in the format (xxx) xxx-xxxx

    phone_pattern = r'\(\d{3}\) \d{3}-\d{4}'

    # Find all matching phone numbers
    phone_numbers = re.findall(phone_pattern, text)
    # Join the numbers with newlines
    return '\n'.join(phone_numbers)
def get_name(text):
    # Split the text into lines
    lines = text.strip().splitlines()

    # Loop through each line to find the first non-empty line
    for line in lines:
        if line.strip():  # Check if the line is not empty
            # Split the line into words and get the first two
            words = line.split()

            return ' '.join(words[:2])  # Return the first two words as a string

    return None  # Return None if there are no lines with text

def zip_code_r(sentence):
    words = sentence.split(" ")
    if words[-1].isdigit():  # Check if the last word is a number
        return " ".join(words[:-1])  # Remove the last word if it's a number
    return sentence  # Return the original sentence if the last word isn't a number
def extract_states(text):
    all_two_letter_caps = re.findall(r'\b[A-Za-z]{2}\b', text)
    return [s for s in all_two_letter_caps]
def extract_info(data,state):
    results=[]
    # Split the input data into sections
    sections = data.split("------------------------------------------")


    for section in sections:
        # Find names in the section
        name = get_name(section) or ""  # Replace with empty string if not found
        name=name.replace("Goes","")
        # Find addresses in the section
        address = get_address(section) or ""  # Replace with empty string if not found
        # Find phone numbers in the section
        phone_numbers = extract_phone_numbers(section) or []  # Replace with empty list if not found
        # Create a dictionary for the current section
        if isinstance(phone_numbers, list):
            phone_numbers='\n'.join(phone_numbers)

        if name and address:  # Ensure both name and phone numbers are presentt
            if extract_states(address.strip())[-1].lower() == state.lower():

                results.append({

                    "Name": name,
                    "Current Address": address,
                    "Phone Numbers":phone_numbers  # Join multiple phone numbers into a single string
                })

    return results


def save_neighbors_infos(results, output_path):
    """Save results to Excel without using pandas"""
    filename = f"{time.time()}.xlsx"
    file_path = os.path.join(output_path, filename)

    # Create workbook and sheet
    if os.path.exists(file_path):
        # Load existing workbook
        wb = load_workbook(file_path)
        ws = wb.active
        start_row = ws.max_row + 1  # Start after last row
    else:
        # Create new workbook
        wb = Workbook()
        ws = wb.active
        # Write headers
        ws.append(["Name", "Current Address", "Phone Numbers"])
        start_row = 2  # Start after header

    # Write data
    for result in results:
        if isinstance(result, dict):  # Handle both single dict and list of dicts
            row = [
                result.get("Name", ""),
                result.get("Current Address", ""),
                result.get("Phone Numbers", "")
            ]
            ws.append(row)

    # Save the file
    wb.save(file_path)
    print(f"Saved data to {file_path}")


def scrape_data(address):
    if not address or "#" in address:
        return


    results = []

    h="Request failed. You will not be charged for this request. Please make sure your url is correct and try again. Protected domains may require adding premium=true OR ultra_premium=true parameter to your request."
    while(h=="Request failed. You will not be charged for this request. Please make sure your url is correct and try again. Protected domains may require adding premium=true OR ultra_premium=true parameter to your request."):
        l=split_address(address)
        if l:
            number, street, city, state, zip_code=l
        else:
            return
        url = f'https://www.fastpeoplesearch.com/address/{number}-{street.replace(" ","-")}_{city.replace(" ","-")}-{state}-{zip_code}'
        payload = {'api_key': 'API_KEY', 'url': url, 'ultra_premium': True}
        try:
            r = requests.get('https://api.scraperapi.com/', params=payload)
            time.sleep(3)
        except ConnectTimeout:
            print("The connection to the server timed out. Please try again.")
            return
        except ConnectionError:
            print("There was a problem connecting to the internet. Please check your connection.")
            return
        except ProtocolError:
            print("There was a problem connecting to the internet. Please check your connection.")
            return
        except ReadTimeout:
            print("There was a problem connecting to the internet. Please check your connection.")
            return
        h = r.text
    if "Request failed." in h:
        print(h)
    soup = BeautifulSoup(h, 'html.parser')
    persons = soup.find_all(class_='card-block')
    if persons:
        p=format_page(persons)

        results=extract_info(p,number,results,state)
    return results


def scrape_name(name,city,state):

    h="Request failed. You will not be charged for this request. Please make sure your url is correct and try again. Protected domains may require adding premium=true OR ultra_premium=true parameter to your request."
    while(h=="Request failed. You will not be charged for this request. Please make sure your url is correct and try again. Protected domains may require adding premium=true OR ultra_premium=true parameter to your request."):

        if name and city and state:
            if not(len(state)==2):
                state=get_state_code(state)
        else:
            return
        url = f'https://www.fastpeoplesearch.com/name/{name.replace(" ","-")}_{city.replace(" ","-")}-{state}'
        payload = {'api_key': '4be86b5322a165f65555e3c247617182', 'url': url, 'ultra_premium': True}
        print(url)
        try:
            r = requests.get('https://api.scraperapi.com/', params=payload)
            time.sleep(3)
        except ConnectTimeout:
            print("The connection to the server timed out. Please try again.")
            return
        except ConnectionError:
            print("There was a problem connecting to the internet. Please check your connection.")
            return
        except ProtocolError:
            print("There was a problem connecting to the internet. Please check your connection.")
            return
        except ReadTimeout:
            print("There was a problem connecting to the internet. Please check your connection.")
            return
        h = r.text
    if "Request failed." in h:
        print(h)
    soup = BeautifulSoup(h, 'html.parser')
    persons = soup.find_all(class_='card-block')
    if persons:
        p=format_page(persons)
        results=extract_info(p,state)
    return results
