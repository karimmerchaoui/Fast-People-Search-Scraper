import threading
import concurrent.futures
import customtkinter as ctk
from tkinter import messagebox, filedialog, ttk
import os
import time
import traceback
from FPSscraper import scrape_name
import random
import re
# =============== GLOBAL VARIABLES ===============
root = None
cancel_event = threading.Event()
loading_active = False
output_folder = ""
input_file = ""
progress_var = None
completed_addresses = 0
total_addresses = 0
results_filename = ""
all_results = []
name_items = {}  # Dictionary to store treeview items
processing_colors = {}  # Dictionary to store processing colors

# GUI Widgets
label_input = None
btn_browse_input = None
label_output = None
btn_browse_output = None
label_loading = None
progress_bar = None
btn_start = None
btn_cancel = None
text_log = None
tree = None  # Treeview widget
label_progress_counter = None
# =============== GUI SETUP ===============
def setup_gui():
    global root, progress_var, label_input, btn_browse_input
    global label_output, btn_browse_output, label_loading, progress_bar
    global btn_start, btn_cancel, text_log, tree

    root = ctk.CTk()
    ctk.set_appearance_mode("Dark")
    root.title("Home Owner Scraper")
    root.geometry("900x630")  # Increased size for treeview

    progress_var = ctk.DoubleVar()

    # Input and Output File Selection (side by side & centered)
    frame_io = ctk.CTkFrame(root, fg_color="transparent")
    frame_io.pack(pady=10)

    # Input frame
    frame_input = ctk.CTkFrame(frame_io, fg_color="transparent", width=400)
    frame_input.pack(side="left", padx=10)
    frame_input.pack_propagate(False)

    label_input = ctk.CTkLabel(frame_input, text="Input File: Not selected")
    label_input.grid(row=0, column=0, padx=5, pady=5)

    btn_browse_input = ctk.CTkButton(frame_input, text="Browse Input", command=browse_input_file)
    btn_browse_input.grid(row=0, column=1, padx=5, pady=5)

    # Output frame
    frame_output = ctk.CTkFrame(frame_io, fg_color="transparent", width=400)
    frame_output.pack(side="left", padx=10)
    frame_output.pack_propagate(False)

    label_output = ctk.CTkLabel(frame_output, text="Output Folder: Not selected")
    label_output.grid(row=0, column=0, padx=5, pady=5)

    btn_browse_output = ctk.CTkButton(frame_output, text="Browse Output", command=browse_output_folder)
    btn_browse_output.grid(row=0, column=1, padx=5, pady=5)

    # Treeview Frame
    tree_frame = ctk.CTkFrame(root)
    tree_frame.pack(pady=10, padx=10, fill="both", expand=True)

    # Create Treeview with dark theme
    style = ttk.Style()
    style.theme_use('default')
    style.configure("Treeview",
                   background="#2a2d2e",
                   foreground="white",
                   rowheight=25,
                   fieldbackground="#2a2d2e",
                   bordercolor="#3b3b3b",
                   borderwidth=0)
    style.map('Treeview', background=[('selected', '#22559b')])
    style.configure("Treeview.Heading",
                   background="#3b3b3b",
                   foreground="white",
                   relief="flat")
    style.map("Treeview.Heading",
              background=[('active', '#4a4a4a')])

    tree = ttk.Treeview(tree_frame, columns=("Status", "Count"), show="tree headings")
    tree.heading("#0", text="Names")
    tree.heading("Status", text="Status")
    tree.heading("Count", text="Found")
    tree.column("#0", width=400)
    tree.column("Status", width=100)
    tree.column("Count", width=60, anchor='center')  # Center the count
    tree.column("Status", width=60, anchor='center')  # Center the count
    tree.column("#0", width=60, anchor='center')  # Center the count

    # Configure tags for different statuses
    tree.tag_configure('pending', foreground='gray')
    tree.tag_configure('processing', foreground='#4fc3f7')  # Light blue
    tree.tag_configure('success', foreground='#81c784')    # Light green
    tree.tag_configure('failed', foreground='#ff8a65')     # Light orange
    tree.tag_configure('error', foreground='#e57373')      # Light red

    # Add scrollbar
    scrollbar = ttk.Scrollbar(tree_frame, orient="vertical", command=tree.yview)
    tree.configure(yscrollcommand=scrollbar.set)
    scrollbar.pack(side="right", fill="y")
    tree.pack(fill="both", expand=True)

    # Progress Frame
    frame_progress = ctk.CTkFrame(root, fg_color="transparent")
    frame_progress.pack(pady=10)

    label_loading = ctk.CTkLabel(frame_progress, text="", font=("Arial", 14))
    label_loading.pack()

    global label_progress_counter
    label_progress_counter = ctk.CTkLabel(root, text="0 / 0 Processed", font=("Arial", 12))
    label_progress_counter.pack(pady=(0, 10))
    progress_bar = ttk.Progressbar(
        root,
        variable=progress_var,
        maximum=100,
        length=650,
        style="blue.Horizontal.TProgressbar"
    )
    progress_bar.pack(pady=5)

    # Action Buttons
    frame_buttons = ctk.CTkFrame(root, fg_color="transparent")
    frame_buttons.pack(pady=10)

    btn_start = ctk.CTkButton(frame_buttons, text="Start Scraping", command=start_scraping)
    btn_start.grid(row=0, column=0, padx=10)

    btn_cancel = ctk.CTkButton(frame_buttons, text="Cancel", command=cancel_operation, state="disabled")
    btn_cancel.grid(row=0, column=1, padx=10)

    # Results Log
    text_log = ctk.CTkTextbox(root, width=850, height=150, state="disabled")
    text_log.pack(pady=10)

    return root

def update_tree_item(item_id, status, color=None, count=None):
    """Update a treeview item with status, color, and count"""
    values = (status.capitalize(), str(count) if count is not None else "")
    tree.item(item_id, values=values, tags=(status.lower(),))
    root.update_idletasks()

# =============== CORE FUNCTIONS ===============


def process_address_batch(address_batch):
    """Process a batch of 10 addresses with simultaneous requests in one thread"""
    global completed_addresses, all_results

    batch_results = []
    threads = []
    results_lock = threading.Lock()

    # Generate a unique color for this batch
    batch_color = "#{:02x}{:02x}{:02x}".format(
        random.randint(100, 200),
        random.randint(100, 200),
        random.randint(100, 200)
    )

    def process_single_address(address):
        nonlocal batch_results
        try:
            if cancel_event.is_set():
                return

            try:
                name, city, state = address
            except ValueError:
                name, city, state = address, "", ""  # Fallback if format is unexpected

            # Update treeview to show processing
            if name in name_items:
                root.after(0, lambda: update_tree_item(name_items[name], "Processing", batch_color))

            result = scrape_name(name, city, state)
            with results_lock:
                if result:
                    url = f'https://www.fastpeoplesearch.com/name/{name.replace(" ", "-")}_{city.replace(" ", "-")}-{state}'
                    log_message(f"✓ Success: {url} (Found {len(result)} records)")


                    batch_results.extend(result)
                    root.after(0, lambda: update_tree_item(
                        name_items[name],
                        "Success",
                        batch_color,
                        len(result)  # Pass the count here
                    ))
                else:
                    empty_result = {'Name': f'{name}', 'Current Address': '', 'Phone Numbers': ''}
                    batch_results.append(empty_result)
                    url = f'https://www.fastpeoplesearch.com/name/{name.replace(" ", "-")}_{city.replace(" ", "-")}-{state}'
                    log_message(f"✗ No data found: {url}")
                    if name in name_items:
                        root.after(0, lambda: update_tree_item(
                            name_items[name],
                            "Failed",
                            None,
                            0  # Show 0 for no results
                        ))

                # Add empty line after each address result
                batch_results.append({'Name': '', 'Current Address': '', 'Phone Numbers': ''})


        except Exception as e:

            error_msg = f"ERROR processing {name}:\n{traceback.format_exc()}"

            log_message(error_msg)

            with results_lock:

                batch_results.append({'Name': f'{name}', 'Current Address': '', 'Phone Numbers': ''})

                batch_results.append({'Name': '', 'Current Address': '', 'Phone Numbers': ''})

                if name in name_items:
                    root.after(0, lambda: update_tree_item(

                        name_items[name],

                        "Error",

                        None,

                        0  # Show 0 for errors

                    ))
        finally:
            global completed_addresses
            completed_addresses += 1
            progress_var.set((completed_addresses / total_addresses) * 100)
            label_progress_counter.configure(text=f"{completed_addresses} / {total_addresses} Processed")

    # Start all 10 requests simultaneously
    for address in address_batch:
        t = threading.Thread(target=process_single_address, args=(address,))
        t.start()
        threads.append(t)

    # Wait for all threads to complete
    for t in threads:
        t.join()

    return batch_results


# =============== CORE FUNCTIONS ===============
def browse_input_file():
    global input_file
    try:
        filetypes = [("Excel files", "*.xlsx;*.xls")]
        filename = filedialog.askopenfilename(title="Select Excel File", filetypes=filetypes)
        if filename:
            input_file = filename
            label_input.configure(text=f"Input File: {os.path.basename(filename)}")
            preview_excel(filename)
    except Exception as e:
        error_msg = f"ERROR in browse_input_file"
        traceback.print_exc()
        messagebox.showerror("Error", error_msg)


def browse_output_folder():
    global output_folder
    try:
        folder = filedialog.askdirectory()
        if folder:
            output_folder = folder
            label_output.configure(text=f"Output Folder: {folder}")
    except Exception as e:
        error_msg = f"ERROR in browse_output_folder"
        traceback.print_exc()
        messagebox.showerror("Error", error_msg)


from openpyxl import Workbook, load_workbook


def preview_excel(filepath):
    try:
        # Clear existing preview
        tree.delete(*tree.get_children())
        name_items.clear()

        # Read Excel file
        wb = load_workbook(filename=filepath)
        ws = wb.active

        # Get names from first column (skip header if exists)
        names = []
        for row in ws.iter_rows(min_row=1, max_col=1, values_only=True):
            if row[0]:  # Only add non-empty names
                names.append(str(row[0]))

        # Populate treeview
        for name in names[:100]:  # Show first 100 as preview
            item_id = tree.insert("", "end", text=name, values=("Pending", ""))  # Empty count initially
            name_items[name] = item_id

        if len(names) > 100:
            log_message(f"Showing first 100/{len(names)} names in preview")

    except Exception as e:
        error_msg = f"ERROR in preview_excel"
        traceback.print_exc()
        messagebox.showerror("Error", error_msg)


def save_all_results(results, output_path):
    """Save results to Excel without pandas"""
    global results_filename

    if not results_filename:
        results_filename = f"results_{int(time.time())}.xlsx"

    file_path = os.path.join(output_path, results_filename)

    try:
        # Create new workbook or load existing
        if os.path.exists(file_path):
            wb = load_workbook(file_path)
            ws = wb.active
            start_row = ws.max_row + 1
        else:
            wb = Workbook()
            ws = wb.active
            # Write headers
            ws.append(["Name", "Current Address", "Phone Numbers"])
            start_row = 2

        # Write data
        for result in results:
            if isinstance(result, dict):
                ws.append([
                    result.get("Name", ""),
                    result.get("Current Address", ""),
                    result.get("Phone Numbers", "")
                ])

        # Save file
        wb.save(file_path)
        log_message(f"Saved data to {file_path}")
        return True

    except Exception as e:
        error_msg = f"ERROR in save_all_results: {str(e)}"
        log_message(error_msg)
        traceback.print_exc()
        return False


def log_message(message):
    try:
        text_log.configure(state="normal")
        text_log.insert("end", f"{message}\n")
        text_log.configure(state="disabled")
        text_log.see("end")
    except Exception as e:
        error_msg = f"ERROR in log_message"
        messagebox.showerror("Error", error_msg)


def clear_log():
    try:
        text_log.configure(state="normal")
        text_log.delete("1.0", "end")
        text_log.configure(state="disabled")
    except Exception as e:
        error_msg = f"ERROR in clear_log"
        traceback.print_exc()
        messagebox.showerror("Error", error_msg)


def loading_animation():
    global loading_active
    try:
        dots = ["⌛ Processing", "⌛ Processing.", "⌛ Processing..", "⌛ Processing..."]
        i = 0

        while loading_active:
            label_loading.configure(text=dots[i % len(dots)])
            i += 1
            time.sleep(0.3)
    except Exception as e:
        error_msg = f"ERROR in loading_animation"
        traceback.print_exc()
        messagebox.showerror("Error", error_msg)


def start_loading():
    global loading_active
    try:
        loading_active = True
        threading.Thread(target=loading_animation, daemon=True).start()
    except Exception as e:
        error_msg = f"ERROR in start_loading"
        traceback.print_exc()
        messagebox.showerror("Error", error_msg)


def stop_loading():
    global loading_active
    try:
        loading_active = False
        label_loading.configure(text="✅ Done!")
    except Exception as e:
        error_msg = f"ERROR in stop_loading"
        traceback.print_exc()
        messagebox.showerror("Error", error_msg)


def cancel_operation():
    try:
        cancel_event.set()
        log_message("Cancellation requested...")
    except Exception as e:
        error_msg = f"ERROR in cancel_operation"
        traceback.print_exc()
        messagebox.showerror("Error", error_msg)




def process_address(address):
    global completed_addresses, all_results

    try:
        if cancel_event.is_set():
            return None

        try:
            name, city, state = address
        except ValueError:
            name, city, state = address, "", ""  # Fallback if format is unexpected

        # Get or create color for this processing thread
        thread_name = threading.current_thread().name

        color = processing_colors[thread_name]

        # Update treeview to show processing - must do this in main thread
        if name in name_items:
            root.after(0, lambda: update_tree_item(name_items[name], "processing", color))

        result = scrape_name(name, city, state)

        if result:
            url = f'https://www.fastpeoplesearch.com/name/{name.replace(" ", "-")}_{city.replace(" ", "-")}-{state}'
            log_message(f"✓ Success: {url}")
            all_results.extend(result)
            update_tree_item(name_items[name],"Success","green")

        else:
            empty_result = [{'Name': '', 'Current Address': f'{name}', 'Phone Numbers': ''}]
            all_results.extend(empty_result)
            url = f'https://www.fastpeoplesearch.com/name/{name.replace(" ", "-")}_{city.replace(" ", "-")}-{state}'
            log_message(f"✗ No data found: {url}")
            update_tree_item(name_items[name], "Failed", "red")

        all_results.extend({'Name': '', 'Current Address': '', 'Phone Numbers': ''})
        return result

    except Exception as e:
        error_msg = f"ERROR processing {name}:\n{traceback.format_exc()}"
        log_message(error_msg)
        all_results.append({'Name': '', 'Current Address': f'{name}', 'Phone Numbers': ''})

        update_tree_item(name_items[name], "Error", "Orange")
        return None
    finally:
        completed_addresses += 1
        progress_var.set((completed_addresses / total_addresses) * 100)
        label_progress_counter.configure(text=f"{completed_addresses} / {total_addresses} Processed")


def start_scraping():
    global completed_addresses, total_addresses
    try:
        if not input_file:
            messagebox.showerror("Error", "Please select an input Excel file first!")
            return

        if not output_folder:
            messagebox.showerror("Error", "Please select an output folder first!")
            return

        # Read addresses from Excel using openpyxl
        workbook = load_workbook(input_file, read_only=True, data_only=True)
        sheet = workbook.active

        names, cities, states = [], [], []

        for row in sheet.iter_rows(min_row=1, max_col=3):  # assuming the first row is a header
            name = str(row[0].value).strip() if row[0].value is not None else ''
            city = str(row[1].value).strip() if row[1].value is not None else ''
            state = str(row[2].value).strip() if row[2].value is not None else ''
            names.append(name)
            cities.append(city)
            states.append(state)

        addresses = list(zip(names, cities, states))
        total_addresses = len(addresses)

        if not addresses:
            messagebox.showerror("Error", "No addresses found in the Excel file!")
            return

        # Reset UI state
        clear_log()
        cancel_event.clear()
        completed_addresses = 0
        progress_var.set(0)

        # Disable controls during operation
        btn_start.configure(state="disabled")
        btn_cancel.configure(state="normal")
        start_loading()

        # Start processing in background thread
        threading.Thread(
            target=run_scraping_tasks,
            args=(addresses,),
            daemon=True
        ).start()

    except Exception as e:
        error_msg = f"ERROR in start_scraping: {e}"
        traceback.print_exc()
        messagebox.showerror("Error", error_msg)




def run_scraping_tasks(addresses):
    global all_results, results_filename

    try:
        # Reset results storage
        all_results = []
        results_filename = f"results_{int(time.time())}.xlsx"

        # Split addresses into batches of 10
        batch_size = 7
        address_batches = [addresses[i:i + batch_size] for i in range(0, len(addresses), batch_size)]

        with concurrent.futures.ThreadPoolExecutor(max_workers=6) as executor:
            # Submit batches to process 10 leads simultaneously in each thread
            futures = [executor.submit(process_address_batch, batch) for batch in address_batches if batch]

            for future in concurrent.futures.as_completed(futures):
                if cancel_event.is_set():
                    executor.shutdown(wait=False)
                    break

                # Collect results from each batch
                batch_result = future.result()
                if batch_result:
                    all_results.extend(batch_result)
        # Save all results at the end
        if all_results:
            # Remove the last empty line if present
            if all_results[-1] == {'Name': '', 'Current Address': '', 'Phone Numbers': ''}:
                all_results = all_results[:-1]

            save_all_results(all_results, output_folder)
            log_message(f"All results saved to: {results_filename}")

    except Exception as e:
        error_msg = f"ERROR in run_scraping_tasks:\n{traceback.format_exc()}"
        log_message(error_msg)
    finally:
        stop_loading()
        btn_start.configure(state="normal")
        btn_cancel.configure(state="disabled")



# =============== MAIN EXECUTION ===============
def main():
    global root
    try:
        root = setup_gui()
        root.mainloop()
    except Exception as e:
        error_msg = f"FATAL ERROR in main:\n{traceback.format_exc()}"
        traceback.print_exc()
        messagebox.showerror("Critical Error", error_msg)
        if root:
            root.destroy()

import sys
def add_manifest():
    if sys.platform == 'win32':
        manifest = """
        <!-- Paste your entire manifest XML here -->
        """
        try:
            import win32api
            import win32con
            win32api.SetFileAttributes("main.py", win32con.FILE_ATTRIBUTE_NORMAL)
            with open("main.py.manifest", "w") as f:
                f.write(manifest.strip())
        except ImportError:
            # Fallback if pywin32 not installed
            with open("main.py.manifest", "w") as f:
                f.write(manifest.strip())


if __name__ == "__main__":

    main()

